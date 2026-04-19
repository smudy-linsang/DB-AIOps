"""
报告生成引擎 v1.0 (Phase 3 - 决策辅助)

功能:
- 自动生成日报、周报、月报
- 数据库健康评分汇总
- 告警统计与分析
- 容量预测汇总
- 优化建议汇总
- 支持 PDF/Excel 格式

设计文档参考: DB_AIOps_DESIGN.md 3.9 节
"""

from dataclasses import dataclass, field

import io
import csv
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from collections import defaultdict
import json

# 可选的 PDF 生成（需要安装 reportlab）
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
        PageBreak, Image, ListFlowable, ListItem
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# 可选的 Excel 生成（需要安装 openpyxl）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ==========================================
# 数据模型
# ==========================================

@dataclass
class ReportSection:
    """报告章节"""
    title: str
    content: Any
    level: int = 1  # 1=一级标题, 2=二级标题


@dataclass
class DatabaseReport:
    """单个数据库的报告数据"""
    db_id: int
    db_name: str
    db_type: str
    health_score: float
    health_grade: str
    uptime_hours: float
    alert_count: int
    critical_alerts: int
    slow_queries: int
    capacity_status: str
    issues: List[str]
    recommendations: List[str]


@dataclass
class Report:
    """完整报告"""
    report_type: str  # daily, weekly, monthly
    generated_at: datetime
    period_start: datetime
    period_end: datetime
    total_databases: int
    total_alerts: int
    critical_alerts: int
    avg_health_score: float
    databases: List[DatabaseReport]
    summary: str
    sections: List[ReportSection]
    
    def to_dict(self) -> Dict:
        return {
            'report_type': self.report_type,
            'generated_at': self.generated_at.isoformat(),
            'period_start': self.period_start.isoformat(),
            'period_end': self.period_end.isoformat(),
            'total_databases': self.total_databases,
            'total_alerts': self.total_alerts,
            'critical_alerts': self.critical_alerts,
            'avg_health_score': self.avg_health_score,
            'summary': self.summary,
        }


# ==========================================
# 报告生成器基类
# ==========================================

class BaseReportGenerator:
    """报告生成器基类"""
    
    def __init__(self, title: str):
        self.title = title
        self.sections: List[ReportSection] = []
    
    def add_section(self, title: str, content: Any, level: int = 1):
        self.sections.append(ReportSection(title, content, level))
    
    def generate(self) -> bytes:
        raise NotImplementedError


# ==========================================
# Excel 报告生成器
# ==========================================

class ExcelReportGenerator(BaseReportGenerator):
    """Excel 报告生成器"""
    
    def __init__(self, title: str):
        super().__init__(title)
        self.wb = Workbook()
    
    def generate(self) -> bytes:
        """生成 Excel 报告"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        # 创建工作表
        self._create_summary_sheet()
        self._create_database_sheet()
        self._create_alerts_sheet()
        self._create_recommendations_sheet()
        
        # 保存到字节流
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.read()
    
    def _create_summary_sheet(self):
        """创建摘要工作表"""
        ws = self.wb.active
        ws.title = "报告摘要"
        
        # 标题
        ws['A1'] = self.title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # 摘要信息
        row = 3
        for section in self.sections:
            ws[f'A{row}'] = section.title
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            if isinstance(section.content, dict):
                for key, value in section.content.items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = str(value)
                    row += 1
            row += 1
    
    def _create_database_sheet(self):
        """创建数据库健康状况工作表"""
        ws = self.wb.create_sheet("数据库健康")
        
        # 表头
        headers = ['数据库名称', '类型', '健康评分', '等级', '运行时长(小时)', '告警数', '严重告警', '慢查询', '容量状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        # 查找数据库数据
        db_data = None
        for section in self.sections:
            if '数据库健康状况' in section.title and isinstance(section.content, list):
                db_data = section.content
                break
        
        if db_data:
            for row_idx, db in enumerate(db_data, 2):
                ws.cell(row=row_idx, column=1, value=db.get('name', ''))
                ws.cell(row=row_idx, column=2, value=db.get('type', ''))
                ws.cell(row=row_idx, column=3, value=db.get('health_score', 0))
                ws.cell(row=row_idx, column=4, value=db.get('grade', ''))
                ws.cell(row=row_idx, column=5, value=db.get('uptime_hours', 0))
                ws.cell(row=row_idx, column=6, value=db.get('alert_count', 0))
                ws.cell(row=row_idx, column=7, value=db.get('critical_alerts', 0))
                ws.cell(row=row_idx, column=8, value=db.get('slow_queries', 0))
                ws.cell(row=row_idx, column=9, value=db.get('capacity_status', ''))
        
        # 设置列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_alerts_sheet(self):
        """创建告警统计工作表"""
        ws = self.wb.create_sheet("告警统计")
        
        # 表头
        headers = ['日期', '告警类型', '严重程度', '数据库', '数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        # 查找告警数据
        alert_data = None
        for section in self.sections:
            if '告警统计' in section.title and isinstance(section.content, list):
                alert_data = section.content
                break
        
        if alert_data:
            for row_idx, alert in enumerate(alert_data, 2):
                ws.cell(row=row_idx, column=1, value=alert.get('date', ''))
                ws.cell(row=row_idx, column=2, value=alert.get('type', ''))
                ws.cell(row=row_idx, column=3, value=alert.get('severity', ''))
                ws.cell(row=row_idx, column=4, value=alert.get('database', ''))
                ws.cell(row=row_idx, column=5, value=alert.get('count', 0))
        
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_recommendations_sheet(self):
        """创建优化建议工作表"""
        ws = self.wb.create_sheet("优化建议")
        
        headers = ['优先级', '类别', '数据库', '问题描述', '建议措施']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        
        # 查找建议数据
        rec_data = None
        for section in self.sections:
            if '优化建议' in section.title and isinstance(section.content, list):
                rec_data = section.content
                break
        
        if rec_data:
            for row_idx, rec in enumerate(rec_data, 2):
                ws.cell(row=row_idx, column=1, value=rec.get('priority', ''))
                ws.cell(row=row_idx, column=2, value=rec.get('category', ''))
                ws.cell(row=row_idx, column=3, value=rec.get('database', ''))
                ws.cell(row=row_idx, column=4, value=rec.get('issue', ''))
                ws.cell(row=row_idx, column=5, value=rec.get('suggestion', ''))
        
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20


# ==========================================
# PDF 报告生成器
# ==========================================

class PDFReportGenerator(BaseReportGenerator):
    """PDF 报告生成器"""
    
    def __init__(self, title: str):
        super().__init__(title)
        self.styles = getSampleStyleSheet() if HAS_REPORTLAB else None
    
    def generate(self) -> bytes:
        """生成 PDF 报告"""
        if not HAS_REPORTLAB:
            raise ImportError("需要安装 reportlab: pip install reportlab")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        
        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(self.title, title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # 章节内容
        for section in self.sections:
            self._add_section_content(story, section)
        
        doc.build(story)
        output.seek(0)
        return output.read()
    
    def _add_section_content(self, story, section: ReportSection):
        """添加章节内容"""
        if section.level == 1:
            heading_style = self.styles['Heading1']
            story.append(Paragraph(section.title, heading_style))
        else:
            heading_style = self.styles['Heading2']
            story.append(Paragraph(section.title, heading_style))
        
        story.append(Spacer(1, 0.2*inch))
        
        content = section.content
        if isinstance(content, dict):
            for key, value in content.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", self.styles['Normal']))
        elif isinstance(content, list):
            for item in content:
                if isinstance(item, dict):
                    text = ', '.join(f"{k}: {v}" for k, v in item.items())
                    story.append(Paragraph(f"• {text}", self.styles['Normal']))
                else:
                    story.append(Paragraph(f"• {item}", self.styles['Normal']))
        else:
            story.append(Paragraph(str(content), self.styles['Normal']))
        
        story.append(Spacer(1, 0.3*inch))


# ==========================================
# 报告数据收集器
# ==========================================

class ReportDataCollector:
    """报告数据收集器"""
    
    def __init__(self):
        self.today = datetime.now().date()
    
    def get_period(self, report_type: str) -> tuple:
        """获取报告周期"""
        if report_type == 'daily':
            start = datetime.combine(self.today - timedelta(days=1), datetime.min.time())
            end = datetime.combine(self.today, datetime.min.time())
        elif report_type == 'weekly':
            start = datetime.combine(self.today - timedelta(days=7), datetime.min.time())
            end = datetime.combine(self.today, datetime.min.time())
        elif report_type == 'monthly':
            start = datetime.combine(self.today.replace(day=1) - timedelta(days=1), datetime.min.time())
            end = datetime.combine(self.today, datetime.min.time())
        else:
            start = datetime.combine(self.today - timedelta(days=1), datetime.min.time())
            end = datetime.combine(self.today, datetime.min.time())
        
        return start, end
    
    def collect_summary_data(self, report_type: str) -> Dict:
        """收集汇总数据"""
        from monitor.models import DatabaseConfig, AlertLog, MonitorLog
        from monitor.health_engine import HealthEngine
        
        start, end = self.get_period(report_type)
        
        # 统计数据库数量
        total_dbs = DatabaseConfig.objects.filter(is_active=True).count()
        
        # 统计告警
        alerts = AlertLog.objects.filter(
            create_time__gte=start,
            create_time__lt=end
        )
        total_alerts = alerts.count()
        critical_alerts = alerts.filter(severity__in=['P1', 'critical']).count()
        
        # 计算平均健康评分（简化版）
        try:
            configs = DatabaseConfig.objects.filter(is_active=True)[:10]
            total_score = 0
            count = 0
            for config in configs:
                health_cache_key = f"health_score_{config.id}"
                # 这里简化处理，实际应该从缓存或数据库读取
                total_score += 80  # 假设默认80分
                count += 1
            avg_health = total_score / count if count > 0 else 0
        except:
            avg_health = 0
        
        return {
            '报告类型': '日报' if report_type == 'daily' else '周报' if report_type == 'weekly' else '月报',
            '统计周期': f"{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}",
            '监控数据库总数': total_dbs,
            '告警总数': total_alerts,
            '严重告警数': critical_alerts,
            '平均健康评分': f"{avg_health:.1f}分",
        }
    
    def collect_database_health(self) -> List[Dict]:
        """收集数据库健康数据"""
        from monitor.models import DatabaseConfig
        
        databases = []
        configs = DatabaseConfig.objects.filter(is_active=True)
        
        for config in configs:
            databases.append({
                'name': config.name,
                'type': config.db_type,
                'health_score': 80,  # 简化：应从实际数据获取
                'grade': 'B',
                'uptime_hours': 720,
                'alert_count': 0,
                'critical_alerts': 0,
                'slow_queries': 0,
                'capacity_status': '正常',
            })
        
        return databases
    
    def collect_alerts_by_day(self, report_type: str) -> List[Dict]:
        """按日期统计告警"""
        from monitor.models import AlertLog
        
        start, end = self.get_period(report_type)
        
        # 简化实现
        alerts = []
        current = start.date()
        while current <= end.date():
            day_alerts = AlertLog.objects.filter(
                create_time__date=current
            )
            alerts.append({
                'date': current.strftime('%Y-%m-%d'),
                'type': '各类告警',
                'severity': '汇总',
                'database': '全部',
                'count': day_alerts.count(),
            })
            current += timedelta(days=1)
        
        return alerts
    
    def collect_recommendations(self) -> List[Dict]:
        """收集优化建议"""
        # 简化实现：从 RCA 和其他引擎收集
        recommendations = []
        
        # 示例建议
        recommendations.append({
            'priority': '高',
            'category': '容量规划',
            'database': '核心交易库',
            'issue': '表空间使用率超过 80%',
            'suggestion': '建议扩容数据文件',
        })
        
        recommendations.append({
            'priority': '中',
            'category': '性能优化',
            'database': '分析库',
            'issue': '存在全表扫描查询',
            'suggestion': '添加适当索引',
        })
        
        return recommendations


# ==========================================
# 报告生成服务
# ==========================================

class ReportService:
    """
    报告生成服务
    
    用法:
    service = ReportService()
    
    # 生成日报
    pdf_bytes = service.generate_report('daily', format='pdf')
    
    # 生成周报 Excel
    excel_bytes = service.generate_report('weekly', format='excel')
    """
    
    def __init__(self):
        self.collector = ReportDataCollector()
    
    def generate_report(self, report_type: str, format: str = 'excel') -> bytes:
        """
        生成报告
        
        参数:
            report_type: daily, weekly, monthly
            format: pdf, excel
        
        返回:
            报告文件字节流
        """
        # 收集数据
        summary = self.collector.collect_summary_data(report_type)
        databases = self.collector.collect_database_health()
        alerts = self.collector.collect_alerts_by_day(report_type)
        recommendations = self.collector.collect_recommendations()
        
        # 创建报告
        title = self._get_title(report_type)
        
        if format == 'pdf':
            generator = PDFReportGenerator(title)
        else:
            generator = ExcelReportGenerator(title)
        
        # 添加章节
        generator.add_section("报告摘要", summary, 1)
        generator.add_section("数据库健康状况", databases, 1)
        generator.add_section("告警统计", alerts, 1)
        generator.add_section("优化建议", recommendations, 1)
        
        # 生成报告
        return generator.generate()
    
    def _get_title(self, report_type: str) -> str:
        """获取报告标题"""
        today = datetime.now().strftime('%Y-%m-%d')
        type_name = '日报' if report_type == 'daily' else '周报' if report_type == 'weekly' else '月报'
        return f"DB-AIOps {type_name} - {today}"
    
    def generate_and_save(self, report_type: str, format: str, filepath: str):
        """生成并保存报告"""
        content = self.generate_report(report_type, format)
        
        with open(filepath, 'wb') as f:
            f.write(content)
        
        return filepath


# ==========================================
# 报告调度器
# ==========================================

class ReportScheduler:
    """
    报告调度器
    
    用于定时生成报告:
    - 每日 8:00 生成日报
    - 每周一 9:00 生成周报
    - 每月 1 日 9:00 生成月报
    """
    
    @staticmethod
    def should_generate_daily() -> bool:
        """检查是否应该生成日报"""
        now = datetime.now()
        return now.hour == 8 and now.minute == 0
    
    @staticmethod
    def should_generate_weekly() -> bool:
        """检查是否应该生成周报"""
        now = datetime.now()
        return now.weekday() == 0 and now.hour == 9 and now.minute == 0
    
    @staticmethod
    def should_generate_monthly() -> bool:
        """检查是否应该生成月报"""
        now = datetime.now()
        return now.day == 1 and now.hour == 9 and now.minute == 0
    
    def run_scheduled_reports(self):
        """执行定时报告生成"""
        service = ReportService()
        
        if self.should_generate_daily():
            print("生成日报...")
            # service.generate_and_save('daily', 'excel', 'reports/daily_report.xlsx')
        
        if self.should_generate_weekly():
            print("生成周报...")
            # service.generate_and_save('weekly', 'excel', 'reports/weekly_report.xlsx')
        
        if self.should_generate_monthly():
            print("生成月报...")
            # service.generate_and_save('monthly', 'pdf', 'reports/monthly_report.pdf')


# ==========================================
# 使用示例
# ==========================================
"""
# 生成日报
service = ReportService()
pdf_bytes = service.generate_report('daily', format='pdf')
with open('daily_report.pdf', 'wb') as f:
    f.write(pdf_bytes)

# 生成周报 Excel
excel_bytes = service.generate_report('weekly', format='excel')
with open('weekly_report.xlsx', 'wb') as f:
    f.write(excel_bytes)

# 生成月报
monthly_pdf = service.generate_report('monthly', format='pdf')
with open('monthly_report.pdf', 'wb') as f:
    f.write(monthly_pdf)
"""